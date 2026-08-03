# -*- coding: utf-8 -*-
from pathlib import Path
from openpyxl import load_workbook

root = Path(r"c:\Users\USER\OneDrive\Documentos\Sergio arboleda\Clases_Enel")
t = (root / "index.html").read_text(encoding="utf-8")
a = t.find('"r6": {')
b = t.find('"fase-1": {', a)
chunk = t[a:b]
art_i = t.find('data-reto="r6"')
art_j = t.find('data-reto-enhance="r6"')
art = t[art_i:art_j]

checks = {
    "r6 output html": '"output": "MCP365_P06_Priorizacion_completada.html"' in chunk,
    "no xlsx deliverable in r6": "Priorizacion_completada.xlsx" not in chunk and "Priorizacion_completada.xlsx" not in art,
    "buildPromptR6 wired": 'caseId === "r6"' in t,
    "static excel download": "planillas/MCP365_P06_Base_priorizacion_tareas.xlsx" in t,
    "ref + practice downloads": "priorizacion-ref" in t and "priorizacion-practica" in t,
    "reset defaults": "data-reset-prompt" in t,
    "no modify excel step": "No necesitas modificarlo" in art,
    "checklist 16": chunk.count("r6-c") == 16,
}
for k, v in checks.items():
    print(("OK" if v else "FAIL"), k)

ws = load_workbook(root / "planillas" / "MCP365_P06_Base_priorizacion_tareas.xlsx").active
h = [ws.cell(7, c).value for c in range(1, 9)]
print("excel headers:", h)
n = sum(1 for r in range(8, 18) if ws.cell(r, 1).value)
print("excel tasks:", n)
for name in (
    "MCP365_P06_Priorizacion_completada.html",
    "MCP365_P06_Practica_personal_priorizacion.html",
):
    p = root / "planillas" / name
    print(name, "bytes", p.stat().st_size if p.exists() else "MISSING")
